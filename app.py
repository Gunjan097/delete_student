import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

import customtkinter as ctk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from remover import login, remove_all_students
from batch_reader import read_credentials

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

_COLS   = ("username", "status", "deleted", "note")
_WIDTHS = {"username": 120, "status": 140, "deleted": 90, "note": 380}


class RemoverApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Student Remover")
        self.geometry("900x600")
        self.minsize(700, 440)

        self._credentials: list = []
        self._row_iids: dict = {}
        self._results: list = []
        self._running = False

        # Pause / resume support
        self._pause_event = threading.Event()
        self._pause_event.set()   # not paused initially
        self._paused = False

        # Bounce animation
        self._animating = False
        self._anim_val = 0.0
        self._anim_dir = 1

        self._build_ui()

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # ── Top controls ──────────────────────────────────────────────────
        top = ctk.CTkFrame(self, corner_radius=8)
        top.grid(row=0, column=0, padx=12, pady=(12, 5), sticky="ew")
        top.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(top, text="Import Credentials", width=150,
                      command=self._import).grid(row=0, column=0, padx=(12, 8), pady=10)

        self._cred_lbl = ctk.CTkLabel(top, text="No file selected",
                                       text_color="gray", anchor="w")
        self._cred_lbl.grid(row=0, column=1, padx=(0, 16), pady=10, sticky="ew")

        self._export_btn = ctk.CTkButton(top, text="Export Results", width=130,
                                          state="disabled", command=self._export_results)
        self._export_btn.grid(row=0, column=2, padx=(0, 8), pady=10)

        self._pause_btn = ctk.CTkButton(
            top, text="Pause", width=80,
            fg_color="#e65100", hover_color="#bf360c",
            state="disabled", command=self._toggle_pause,
        )
        self._pause_btn.grid(row=0, column=3, padx=(0, 8), pady=10)

        self._run_btn = ctk.CTkButton(
            top, text="Remove All Students", width=190,
            fg_color="#c62828", hover_color="#8e0000",
            state="disabled", command=self._confirm_run,
        )
        self._run_btn.grid(row=0, column=4, padx=(0, 12), pady=10)

        # ── Results table ──────────────────────────────────────────────────
        tf = ctk.CTkFrame(self, corner_radius=8)
        tf.grid(row=1, column=0, padx=12, pady=5, sticky="nsew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        ttk.Style().configure("Treeview", rowheight=25)
        ttk.Style().configure("Treeview.Heading", font=("", 10, "bold"))

        headers = {"username": "School ID", "status": "Status",
                   "deleted": "Deleted", "note": "Note"}
        self._tree = ttk.Treeview(tf, columns=_COLS, show="headings", selectmode="browse")
        for c in _COLS:
            self._tree.heading(c, text=headers[c])
            self._tree.column(c, width=_WIDTHS[c], anchor="w",
                              minwidth=40, stretch=(c == "note"))
        self._tree.tag_configure("running", foreground="#2196f3")
        self._tree.tag_configure("paused",  foreground="#ff9800")
        self._tree.tag_configure("done",    foreground="#4caf50")
        self._tree.tag_configure("error",   foreground="#e05252")

        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew", padx=(4, 0), pady=(4, 0))
        vsb.grid(row=0, column=1, sticky="ns",  pady=(4, 0))
        hsb.grid(row=1, column=0, sticky="ew",  padx=(4, 0))

        # ── Progress + status ──────────────────────────────────────────────
        self._progress = ctk.CTkProgressBar(self, height=10, corner_radius=5)
        self._progress.set(0)
        self._progress.grid(row=2, column=0, padx=12, pady=(4, 2), sticky="ew")
        self._progress.grid_remove()

        self._status = ctk.CTkLabel(
            self, text="Import a credentials file to begin.",
            anchor="w", text_color="gray", height=26)
        self._status.grid(row=3, column=0, padx=14, pady=(0, 8), sticky="ew")

    # ── Actions ───────────────────────────────────────────────────────────────

    def _import(self):
        path = filedialog.askopenfilename(
            title="Select credentials file",
            filetypes=[("Spreadsheet", "*.xlsx *.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            creds = read_credentials(path)
        except Exception as e:
            messagebox.showerror("Read Error", str(e))
            return
        if not creds:
            messagebox.showwarning("Empty", "No credentials found in the file.")
            return

        self._credentials = creds
        self._cred_lbl.configure(
            text=f"{os.path.basename(path)}  ({len(creds)} school(s))",
            text_color="white")
        self._populate_table()
        self._run_btn.configure(state="normal")
        self._set_status(f"Loaded {len(creds)} school(s). Click 'Remove All Students' to proceed.")

    def _populate_table(self):
        self._tree.delete(*self._tree.get_children())
        self._row_iids.clear()
        self._results.clear()
        for u, _ in self._credentials:
            iid = self._tree.insert("", "end", values=(u, "Pending", "—", ""))
            self._row_iids[u] = iid

    def _confirm_run(self):
        n = len(self._credentials)
        answer = messagebox.askyesno(
            "Confirm Deletion",
            f"You are about to remove ALL students from {n} school(s).\n\n"
            "This action CANNOT be undone.\n\n"
            "Are you absolutely sure?",
            icon="warning",
        )
        if not answer:
            return
        answer2 = messagebox.askyesno(
            "Final Confirmation",
            f"Confirmed — delete all students from {n} school(s)?",
            icon="warning",
        )
        if not answer2:
            return
        self._start_run()

    def _start_run(self):
        self._running = True
        self._paused = False
        self._pause_event.set()
        self._run_btn.configure(state="disabled", text="Running…")
        self._pause_btn.configure(state="normal", text="Pause")
        self._export_btn.configure(state="disabled")
        self._populate_table()
        self._progress.set(0)
        self._progress.grid()
        self._start_bounce()
        self._set_status("Starting…")
        threading.Thread(target=self._worker,
                         args=(list(self._credentials),),
                         daemon=True).start()

    def _toggle_pause(self):
        if not self._running:
            return
        if self._paused:
            # Resume
            self._paused = False
            self._pause_event.set()
            self._pause_btn.configure(text="Pause", fg_color="#e65100", hover_color="#bf360c")
            self._set_status("Resumed…")
            self._start_bounce()
        else:
            # Pause
            self._paused = True
            self._pause_event.clear()
            self._pause_btn.configure(text="Resume", fg_color="#388e3c", hover_color="#1b5e20")
            self._animating = False
            self._set_status("Paused — click Resume to continue.")

    def _worker(self, creds):
        total = len(creds)
        results = []

        for idx, (u, p) in enumerate(creds):
            # Wait here if paused
            self._pause_event.wait()

            self.after(0, self._set_row, u, "Logging in…", "—", "", "running")
            self._set_status_threadsafe(f"Processing {idx + 1} / {total} — {u}")

            try:
                token, school_id = login(u, p)
                self.after(0, self._set_row, u, "Removing…", "—", "", "running")
                result = remove_all_students(token, school_id)

                deleted = _extract_count(result)
                note    = str(result) if deleted is None else ""
                deleted_str = str(deleted) if deleted is not None else "—"

                self.after(0, self._set_row, u, "Done ✓", deleted_str, note, "done")
                results.append({"username": u, "status": "Done",
                                 "deleted": deleted_str, "note": note})
            except Exception as e:
                self.after(0, self._set_row, u, "Error ✗", "—", str(e), "error")
                results.append({"username": u, "status": "Error",
                                 "deleted": "—", "note": str(e)})

            self.after(0, self._progress.set, (idx + 1) / total)

        self._results = results
        done_count  = sum(1 for r in results if r["status"] == "Done")
        error_count = total - done_count
        summary = (f"Complete — {done_count}/{total} succeeded"
                   + (f", {error_count} failed" if error_count else ""))

        self.after(0, self._finish, summary, bool(error_count))

    def _finish(self, summary, has_errors):
        self._running = False
        self._paused = False
        self._pause_event.set()
        self._stop_bounce()
        self._set_status(summary, has_errors)
        self._run_btn.configure(state="normal", text="Remove All Students")
        self._pause_btn.configure(
            state="disabled", text="Pause",
            fg_color="#e65100", hover_color="#bf360c",
        )
        self._export_btn.configure(state="normal")

    def _export_results(self):
        if not self._results:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"deletion_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            return
        try:
            _write_results(self._results, path)
            messagebox.showinfo("Saved", f"Results saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _set_row(self, username, status, deleted, note, tag="running"):
        iid = self._row_iids.get(username)
        if iid:
            self._tree.item(iid, values=(username, status, deleted, note), tags=(tag,))
            self._tree.see(iid)

    def _set_status_threadsafe(self, msg, error=False):
        self.after(0, self._set_status, msg, error)

    def _set_status(self, msg, error=False):
        self._status.configure(text=msg,
                                text_color="#e05252" if error else "gray")

    def _start_bounce(self):
        self._animating = True
        self._anim_val, self._anim_dir = 0.0, 1
        self._progress.grid()
        self._bounce_tick()

    def _bounce_tick(self):
        if not self._animating:
            return
        self._anim_val += 0.025 * self._anim_dir
        if self._anim_val >= 1.0:
            self._anim_val, self._anim_dir = 1.0, -1
        elif self._anim_val <= 0.0:
            self._anim_val, self._anim_dir = 0.0, 1
        self._progress.set(self._anim_val)
        self.after(25, self._bounce_tick)

    def _stop_bounce(self):
        self._animating = False
        self._progress.set(1)
        self.after(500, self._progress.grid_remove)


# ── Utility functions ─────────────────────────────────────────────────────────

def _extract_count(result) -> int | None:
    if isinstance(result, int):
        return result
    if isinstance(result, dict):
        for key in ("deleted", "removed", "count", "total", "studentDeleted",
                    "studentsDeleted", "deletedCount"):
            if key in result:
                try:
                    return int(result[key])
                except (TypeError, ValueError):
                    pass
        for v in result.values():
            try:
                return int(v)
            except (TypeError, ValueError):
                pass
    return None


def _write_results(results: list, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deletion Results"

    fill  = PatternFill("solid", fgColor="8B0000")
    font  = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center")

    headers = ["School ID", "Status", "Students Deleted", "Note", "Timestamp"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill, cell.font, cell.alignment = fill, font, align

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for ri, r in enumerate(results, 2):
        ws.cell(ri, 1, r["username"])
        ws.cell(ri, 2, r["status"])
        ws.cell(ri, 3, r["deleted"])
        ws.cell(ri, 4, r["note"])
        ws.cell(ri, 5, ts)

    for ci in range(1, 6):
        max_len = max(
            len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[chr(64 + ci)].width = min(max_len + 2, 50)

    wb.save(path)
